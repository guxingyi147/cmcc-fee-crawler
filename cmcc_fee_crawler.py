# -*- coding: utf-8 -*-
"""
中国移动资费公示桌面爬取工具
基于 Python + PySide6 + Selenium

功能：
  - 可视化浏览器，用户手动选择筛选条件
  - 自动滚动页面触发懒加载，抓取套餐数据
  - 以方案编号为唯一ID自动去重
  - 支持关键字搜索、价格区间筛选
  - 导出Excel完整保留全部18个字段
"""

import sys
import os
import re
import json
import time
import subprocess
import threading
import zipfile
import urllib.request
import shutil
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QTextEdit, QFileDialog, QSizePolicy, QMessageBox, QAbstractItemView
)
from PySide6.QtCore import Qt, Signal, QObject
from PySide6.QtGui import QFont, QDoubleValidator

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    WebDriverException, StaleElementReferenceException
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================
# 常量
# ============================================================
TARGET_URL = "https://www.10086.cn/fee/"

COLUMNS = [
    "方案编号", "套餐名称", "资费标准", "资费类型", "适用地区",
    "适用范围", "销售渠道", "上线日期", "下线日期", "有效期限",
    "在网要求", "国内通话", "国内通用流量", "定向流量", "短信",
    "退订方式", "违约责任", "其他服务内容"
]
COL_COUNT = len(COLUMNS)

# 列索引
COL_PLAN_ID = 0
COL_NAME = 1
COL_PRICE = 2

# CSS 选择器候选列表（按优先级排序）
CARD_SELECTORS = [
    ".tariff-item-container",     # 中国移动资费公示页面精确选择器
    ".fee-card-item",
    ".tariff-card",
    ".plan-card",
    ".card-item",
]

# 键值对字段映射
FIELD_MAP = {
    "方案编号": COL_PLAN_ID,
    "资费标准": COL_PRICE,
    "资费类型": 3,
    "适用地区": 4,
    "适用范围": 5,
    "销售渠道": 6,
    "上线日期": 7,
    "下线日期": 8,
    "有效期限": 9,
    "在网要求": 10,
    "退订方式": 15,
    "违约责任": 16,
    "其他服务内容": 17,
}

# 排除关键字（用于识别套餐名称）
EXCLUDE_KEYWORDS = [
    "资费标准", "适用地区", "方案编号", "适用范围", "销售渠道",
    "上线日期", "下线日期", "有效期限", "在网要求", "退订方式",
    "违约责任", "其他服务", "国内通话", "国内通用流量", "定向流量",
    "短信", "资费类型",
]


# ============================================================
# 信号
# ============================================================
class WorkerSignals(QObject):
    log = Signal(str)
    status = Signal(str)
    new_data = Signal(list)       # 爬虫线程解析完的数据，发给主线程显示
    type_switched = Signal(str)   # 当前切换到的资费类型名称
    error = Signal(str)
    finished = Signal()


# ============================================================
# 爬虫线程 —— 负责滚动浏览器 + 自动遍历资费类型
# ============================================================

# 资费类型下拉框的 DOM 选择器
TYPE_DROPDOWN_CONTAINER = ".line-3 .select-container.the-select"
TYPE_DROPDOWN_BOX = ".select-box"
TYPE_DROPDOWN_LIST = ".select-list-box"
TYPE_DROPDOWN_ITEM = ".select-item"


class CrawlerThread(threading.Thread):

    def __init__(self, driver, signals, stop_event,
                 traverse_types=True):
        super().__init__(daemon=True)
        self.driver = driver
        self.signals = signals
        self.stop_event = stop_event
        self.traverse_types = traverse_types
        self.seen_ids = set()       # 去重用（线程内维护）

    def run(self):
        try:
            if self.traverse_types:
                self._traverse_types_loop()
            else:
                self._scroll_loop()
        except Exception as e:
            self.signals.error.emit(f"抓取异常: {str(e)}")
            self.signals.log.emit(f"[错误] {str(e)}")
        finally:
            self.signals.finished.emit()

    # ---------- 自动遍历资费类型 ----------

    def _traverse_types_loop(self):
        self.signals.status.emit("遍历资费类型中...")
        self.signals.log.emit(
            "[信息] ========== 开始自动遍历所有资费类型 ==========")

        type_options = self._get_type_options()
        if not type_options:
            self.signals.log.emit(
                "[警告] 未能获取资费类型选项列表，"
                "回退为仅抓取当前页面")
            self._scroll_loop()
            return

        self.signals.log.emit(
            f"[信息] 检测到 {len(type_options)} 个资费类型: "
            + "、".join(t["label"] for t in type_options))

        for idx, type_opt in enumerate(type_options):
            if self.stop_event.is_set():
                break

            type_label = type_opt["label"]
            type_value = type_opt["value"]
            self.signals.log.emit(
                f"\n[信息] ----- 切换资费类型 [{idx+1}/{len(type_options)}]: "
                f"{type_label} -----")
            self.signals.type_switched.emit(
                f"遍历中 [{idx+1}/{len(type_options)}] {type_label}")
            self.signals.status.emit(
                f"正在抓取: {type_label} "
                f"({idx+1}/{len(type_options)})")

            if not self._switch_type(type_value, type_label):
                self.signals.log.emit(
                    f"[警告] 切换资费类型「{type_label}」失败，跳过")
                continue

            self.signals.log.emit("[信息] 等待页面加载...")
            time.sleep(2.0)

            self._scroll_for_current_type()

        self.signals.log.emit(
            "\n[信息] ========== 所有资费类型遍历完毕 ==========")
        self.signals.status.emit("遍历完成")

    def _get_type_options(self):
        try:
            containers = self.driver.find_elements(
                By.CSS_SELECTOR, TYPE_DROPDOWN_CONTAINER)
            if not containers:
                self.signals.log.emit("[警告] 未找到资费类型下拉框容器")
                return []

            type_container = containers[0]
            box = type_container.find_element(
                By.CSS_SELECTOR, TYPE_DROPDOWN_BOX)
            box.click()
            time.sleep(0.8)

            items = type_container.find_elements(
                By.CSS_SELECTOR, TYPE_DROPDOWN_ITEM)
            options = []
            for item in items:
                label_el = item.find_element(
                    By.CSS_SELECTOR, ".item-label")
                value = item.get_attribute("data-value") or ""
                label = label_el.text.strip()
                if label:
                    options.append({"value": value, "label": label})

            self.driver.execute_script(
                "document.querySelector('.select-box')?.click();"
                "document.body.click();")
            time.sleep(0.3)
            return options

        except Exception as e:
            self.signals.log.emit(f"[警告] 获取资费类型选项失败: {e}")
            return []

    def _switch_type(self, value, label):
        try:
            containers = self.driver.find_elements(
                By.CSS_SELECTOR, TYPE_DROPDOWN_CONTAINER)
            if not containers:
                return False

            type_container = containers[0]
            box = type_container.find_element(
                By.CSS_SELECTOR, TYPE_DROPDOWN_BOX)
            box.click()
            time.sleep(0.8)

            items = type_container.find_elements(
                By.CSS_SELECTOR, TYPE_DROPDOWN_ITEM)
            for item in items:
                item_value = item.get_attribute("data-value") or ""
                item_label_el = item.find_element(
                    By.CSS_SELECTOR, ".item-label")
                item_label = item_label_el.text.strip()
                if item_value == value or item_label == label:
                    item.click()
                    self.signals.log.emit(
                        f"[信息] 已选择资费类型: {label}")
                    time.sleep(0.5)
                    return True

            self.signals.log.emit(
                f"[警告] 未找到选项 value={value} label={label}")
            return False

        except Exception as e:
            self.signals.log.emit(f"[警告] 切换资费类型出错: {e}")
            return False

    def _scroll_for_current_type(self):
        """滚动抓取当前资费类型下的全部数据。
        用 JS 快速获取卡片数量判断是否有新数据，有新数据时才做完整解析。
        """
        # 先滚回顶部
        try:
            self.driver.execute_script("window.scrollTo(0, 0)")
            time.sleep(0.8)
        except Exception:
            pass

        # 解析初始页面
        last_card_count = self._get_card_count_js()
        self._parse_all_and_emit()

        no_new_count = 0
        max_no_new = 3
        scroll_round = 0

        while not self.stop_event.is_set():
            scroll_round += 1
            try:
                # 滚到底部
                self.driver.execute_script(
                    "window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(1.0)

                # 用 JS 快速检查卡片数量（毫秒级）
                current_count = self._get_card_count_js()

                if current_count == last_card_count:
                    no_new_count += 1
                    self.signals.log.emit(
                        f"[信息] 卡片数未变化 ({current_count}条, "
                        f"{no_new_count}/{max_no_new})")
                    if no_new_count >= max_no_new:
                        self.signals.log.emit(
                            f"[信息] 资费类型「{self._current_type_label()}」"
                            f"加载完毕（{scroll_round}轮，{current_count}条）")
                        break
                else:
                    no_new_count = 0
                    self.signals.log.emit(
                        f"[信息] 第{scroll_round}轮: "
                        f"{last_card_count} → {current_count} 条")
                    # 只在有新数据时才做完整解析
                    self._parse_all_and_emit()
                    last_card_count = current_count

            except WebDriverException as e:
                self.signals.log.emit(f"[警告] 滚动异常: {e}")
                try:
                    _ = self.driver.current_url
                except Exception:
                    self.signals.error.emit("浏览器已断开连接")
                    return
                time.sleep(3)

    def _get_card_count_js(self):
        """用 JS 快速获取页面上的卡片数量（毫秒级）"""
        try:
            count = self.driver.execute_script("""
                var selectors = [
                    '.tariff-item-container',
                    '.fee-card-item', '.tariff-card', '.plan-card',
                    '.card-item'
                ];
                for (var i = 0; i < selectors.length; i++) {
                    var els = document.querySelectorAll(selectors[i]);
                    if (els.length > 0) return els.length;
                }
                return 0;
            """)
            return count or 0
        except Exception:
            return 0

    def _parse_all_and_emit(self):
        """用 JS 批量提取所有卡片的名称和文本，纯 Python 解析"""
        try:
            # 一次性用 JS 提取所有卡片的名称和文本
            items = self.driver.execute_script("""
                var cards = document.querySelectorAll('.tariff-item-container');
                var result = [];
                for (var i = 0; i < cards.length; i++) {
                    var nameEl = cards[i].querySelector('.item-name');
                    result.push({
                        name: nameEl ? nameEl.textContent.trim() : '',
                        text: cards[i].innerText
                    });
                }
                return result;
            """)
            if not items:
                return

            new_batch = []
            for item in items:
                try:
                    data = parse_text(item["text"], item["name"])
                    if not data:
                        continue
                    plan_id = data[COL_PLAN_ID]
                    if not plan_id:
                        continue
                    if plan_id in self.seen_ids:
                        continue
                    self.seen_ids.add(plan_id)
                    new_batch.append(data)
                except Exception:
                    continue

            if new_batch:
                self.signals.new_data.emit(new_batch)

        except Exception:
            pass

    def _current_type_label(self):
        try:
            containers = self.driver.find_elements(
                By.CSS_SELECTOR, TYPE_DROPDOWN_CONTAINER)
            if containers:
                tips = containers[0].find_element(
                    By.CSS_SELECTOR, ".tipsText")
                return tips.text.strip()
        except Exception:
            pass
        return ""

    # ---------- 单类型滚动（兼容旧逻辑） ----------

    def _scroll_loop(self):
        self.signals.status.emit("抓取中...")
        self.signals.log.emit("[信息] 开始自动滚动页面...")

        last_count = self._get_card_count_js()
        self._parse_all_and_emit()
        stable_count = 0
        max_stable = 5
        scroll_round = 0

        while not self.stop_event.is_set():
            scroll_round += 1
            try:
                self.driver.execute_script(
                    "window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(1.0)

                current_count = self._get_card_count_js()

                if current_count == last_count:
                    stable_count += 1
                    if stable_count >= max_stable:
                        self.signals.log.emit(
                            "[信息] 已滚动到底部，数据加载完毕")
                        break
                else:
                    stable_count = 0
                    self._parse_all_and_emit()
                    last_count = current_count

                self.signals.log.emit(
                    f"[信息] 第 {scroll_round} 轮，"
                    f"累计 {current_count} 条")

            except WebDriverException as e:
                self.signals.log.emit(f"[警告] 浏览器异常: {e}")
                try:
                    _ = self.driver.current_url
                except Exception:
                    self.signals.error.emit("浏览器已断开连接")
                    return
                time.sleep(3)

        self.signals.status.emit("抓取完成")

    # ---------- 通用方法 ----------

    def _page_height(self):
        try:
            return self.driver.execute_script(
                "return document.body.scrollHeight")
        except Exception:
            return 0


# ============================================================
# 数据解析工具函数
# ============================================================
def extract_field_value(text, label):
    """从文本中提取 'label：value' 格式的字段值"""
    pattern = rf'{re.escape(label)}[：:\s]\s*(.+?)(?:\n|$)'
    match = re.search(pattern, text)
    if match:
        value = match.group(1).strip()
        value = re.sub(r'^[●\-\*•·]\s*', '', value)
        if value:
            return value
    return ""


def extract_resource_value(text, label, units):
    """提取资源类字段（通话分钟、流量、短信条数）"""
    # 先尝试键值对
    value = extract_field_value(text, label)
    if value:
        for unit in units:
            if unit in value:
                return value
        # 纯数字补单位
        if re.match(r'^[\d.]+$', value):
            if "通话" in label:
                return f"{value}分钟"
            elif "流量" in label:
                num = float(value)
                return f"{num/1024:.1f}GB" if num >= 1024 else f"{value}MB"
            elif "短信" in label:
                return f"{value}条"

    # 正则提取 label + 数字 + 单位
    unit_pat = "|".join(re.escape(u) for u in units)
    m = re.search(
        rf'{re.escape(label)}[：:\s]*([\d.]+)\s*({unit_pat})',
        text, re.IGNORECASE)
    if m:
        return f"{m.group(1)}{m.group(2)}"

    return ""


def extract_plan_name(card, text):
    """提取套餐名称"""
    # 方法1：标题元素
    for sel in ["h3", "h4", "h5",
                "[class*='title']", "[class*='name']",
                "[class*='plan-name']", "[class*='card-title']"]:
        try:
            for t in card.find_elements(By.CSS_SELECTOR, sel):
                t_text = t.text.strip()
                if (t_text and 2 < len(t_text) < 100
                        and not any(kw in t_text for kw in EXCLUDE_KEYWORDS)):
                    return t_text
        except Exception:
            continue

    # 方法2：第一行非字段文本
    for line in text.strip().split("\n"):
        line = line.strip()
        if (line and 2 < len(line) < 80
                and not line.startswith("●")
                and not any(kw in line for kw in EXCLUDE_KEYWORDS)):
            return line
    return ""


def parse_text(text, plan_name=""):
    """从纯文本解析套餐数据 → 返回长度为 COL_COUNT 的列表，失败返回 None"""
    if not text or len(text) < 10:
        return None

    data = [""] * COL_COUNT

    # 套餐名称：优先使用 JS 提取的精确名称
    if plan_name and len(plan_name) > 1:
        data[COL_NAME] = plan_name
    else:
        # 回退：取第一行非字段文本
        for line in text.strip().split("\n"):
            line = line.strip()
            if (line and 2 < len(line) < 80
                    and not line.startswith("●")
                    and not any(kw in line for kw in EXCLUDE_KEYWORDS)
                    and not re.match(r'^[\d.]+元', line)):
                data[COL_NAME] = line
                break

    # 键值对字段
    for label, col_idx in FIELD_MAP.items():
        val = extract_field_value(text, label)
        if val:
            data[col_idx] = val

    # 4 个特殊字段
    data[11] = extract_resource_value(text, "国内通话", ["分钟", "分"])
    data[12] = extract_resource_value(
        text, "国内通用流量", ["GB", "MB", "gb", "mb"])
    data[13] = extract_resource_value(
        text, "定向流量", ["GB", "MB", "gb", "mb"])
    data[14] = extract_resource_value(text, "短信", ["条"])

    return data


def parse_card(card):
    """解析单个套餐卡片 → 返回长度为 COL_COUNT 的列表，失败返回 None"""
    try:
        text = card.text
    except Exception:
        return None

    if not text or len(text) < 10:
        return None

    data = [""] * COL_COUNT

    # 套餐名称
    data[COL_NAME] = extract_plan_name(card, text)

    # 键值对字段
    for label, col_idx in FIELD_MAP.items():
        val = extract_field_value(text, label)
        if val:
            data[col_idx] = val

    # 4 个特殊字段
    data[11] = extract_resource_value(text, "国内通话", ["分钟", "分"])
    data[12] = extract_resource_value(
        text, "国内通用流量", ["GB", "MB", "gb", "mb"])
    data[13] = extract_resource_value(
        text, "定向流量", ["GB", "MB", "gb", "mb"])
    data[14] = extract_resource_value(text, "短信", ["条"])

    return data


def find_cards(driver):
    """在页面中查找所有套餐卡片元素"""
    # CSS 选择器
    for sel in CARD_SELECTORS:
        try:
            cards = driver.find_elements(By.CSS_SELECTOR, sel)
            if cards:
                return cards
        except Exception:
            continue

    # 回退：通过文本定位父容器
    try:
        elements = driver.find_elements(
            By.XPATH,
            "//*[contains(text(),'方案编号') or contains(text(),'资费标准')]")
        cards, seen = [], set()
        for el in elements:
            try:
                parent = el.find_element(By.XPATH,
                    "./ancestor::*[contains(@class,'card') or "
                    "contains(@class,'item') or contains(@class,'plan') or "
                    "contains(@class,'tariff') or contains(@class,'fee') or "
                    "contains(@class,'list') or contains(@class,'content') or "
                    "contains(@class,'detail')][1]")
                pid = parent.id
                if pid not in seen:
                    seen.add(pid)
                    cards.append(parent)
            except Exception:
                continue
        if cards:
            return cards
    except Exception:
        pass

    return []


# ============================================================
# 主窗口
# ============================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("中国移动资费公示爬取工具")
        self.setMinimumSize(1200, 750)
        self.resize(1400, 850)

        self.driver = None
        self.crawler_thread = None
        self.stop_event = threading.Event()
        self.all_data = []          # 全部数据列表
        self.seen_ids = set()       # 已见方案编号（去重用）
        self.is_crawling = False
        self.signals = WorkerSignals()

        self._connect_signals()
        self._build_ui()
        self.append_log("[信息] 应用已启动，请点击「启动浏览器」开始")

    # ---------- 信号连接 ----------

    def _connect_signals(self):
        self.signals.log.connect(self.append_log)
        self.signals.status.connect(self.update_status)
        self.signals.new_data.connect(self._on_new_data)
        self.signals.type_switched.connect(self._on_type_switched)
        self.signals.error.connect(self._on_error)
        self.signals.finished.connect(self._on_crawl_finished)

    # ---------- UI 构建 ----------

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        # ---- 1. 顶部按钮 ----
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self.btn_launch = QPushButton("启动浏览器")
        self.btn_start = QPushButton("开始抓取")
        self.btn_stop = QPushButton("停止抓取")
        self.btn_clear = QPushButton("清空数据")
        self.btn_export = QPushButton("导出Excel")

        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(False)

        for b in (self.btn_launch, self.btn_start, self.btn_stop,
                  self.btn_clear, self.btn_export):
            b.setFixedHeight(36)
            b.setMinimumWidth(110)
            btn_row.addWidget(b)
        btn_row.addStretch()

        self.btn_launch.clicked.connect(self.launch_browser)
        self.btn_start.clicked.connect(self.start_crawl)
        self.btn_stop.clicked.connect(self.stop_crawl)
        self.btn_clear.clicked.connect(self.clear_data)
        self.btn_export.clicked.connect(self.export_excel)
        root.addLayout(btn_row)

        # ---- 2. 搜索筛选 ----
        filter_row = QHBoxLayout()
        filter_row.setSpacing(12)

        filter_row.addWidget(QLabel("关键字搜索:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("按套餐名称 / 方案编号模糊搜索")
        self.search_input.setFixedWidth(280)
        self.search_input.textChanged.connect(self.apply_filter)
        filter_row.addWidget(self.search_input)

        filter_row.addWidget(QLabel("价格区间:"))
        self.price_min = QLineEdit()
        self.price_min.setPlaceholderText("最低")
        self.price_min.setFixedWidth(80)
        self.price_min.setValidator(QDoubleValidator(0, 99999, 2))
        filter_row.addWidget(self.price_min)

        filter_row.addWidget(QLabel("—"))

        self.price_max = QLineEdit()
        self.price_max.setPlaceholderText("最高")
        self.price_max.setFixedWidth(80)
        self.price_max.setValidator(QDoubleValidator(0, 99999, 2))
        filter_row.addWidget(self.price_max)

        btn_f = QPushButton("筛选")
        btn_f.setFixedWidth(60)
        btn_f.clicked.connect(self.apply_filter)
        filter_row.addWidget(btn_f)

        btn_r = QPushButton("重置")
        btn_r.setFixedWidth(60)
        btn_r.clicked.connect(self.reset_filter)
        filter_row.addWidget(btn_r)

        filter_row.addStretch()

        self.label_count = QLabel("共 0 条数据")
        self.label_count.setStyleSheet(
            "font-weight:bold; color:#1890ff; font-size:13px;")
        filter_row.addWidget(self.label_count)
        root.addLayout(filter_row)

        # ---- 3. 数据表格 ----
        self.table = QTableWidget()
        self.table.setColumnCount(COL_COUNT)
        self.table.setHorizontalHeaderLabels(COLUMNS)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setDefaultSectionSize(28)
        self.table.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        hdr = self.table.horizontalHeader()
        widths = {
            0: 110, 1: 220, 2: 100, 3: 80, 4: 80,
            5: 100, 6: 120, 7: 100, 8: 100, 9: 140,
            10: 80, 11: 100, 12: 110, 13: 100, 14: 80,
            15: 160, 16: 160, 17: 200,
        }
        for c, w in widths.items():
            hdr.resizeSection(c, w)
        root.addWidget(self.table, stretch=1)

        # ---- 4. 底部状态栏 ----
        bot = QVBoxLayout()
        bot.setSpacing(4)

        status_row = QHBoxLayout()
        self.label_status = QLabel("状态: 就绪")
        self.label_status.setStyleSheet(
            "color:#52c41a; font-weight:bold;")
        status_row.addWidget(self.label_status)

        self.label_current_type = QLabel("")
        self.label_current_type.setStyleSheet(
            "color:#722ed1; font-weight:bold; font-size:12px;")
        status_row.addWidget(self.label_current_type)

        status_row.addStretch()
        bot.addLayout(status_row)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(130)
        self.log_text.setStyleSheet(
            "QTextEdit{background:#1e1e1e; color:#d4d4d4;"
            "font-family:'Consolas','Microsoft YaHei',monospace;"
            "font-size:12px; border:1px solid #444; border-radius:4px;}")
        bot.addWidget(self.log_text)
        root.addLayout(bot)

        # ---- 全局样式 ----
        self.setStyleSheet("""
            QMainWindow{background:#f5f5f5;}
            QPushButton{background:#fff; border:1px solid #d9d9d9;
                border-radius:4px; padding:6px 16px; font-size:13px; color:#333;}
            QPushButton:hover{border-color:#40a9ff; color:#40a9ff;}
            QPushButton:pressed{background:#e6f7ff; border-color:#1890ff;}
            QPushButton:disabled{color:#bfbfbf; border-color:#d9d9d9;
                background:#f5f5f5;}
            QLineEdit{border:1px solid #d9d9d9; border-radius:4px;
                padding:5px 10px; font-size:13px; background:#fff;}
            QLineEdit:focus{border-color:#40a9ff;}
            QTableWidget{background:#fff; border:1px solid #e8e8e8;
                border-radius:4px; gridline-color:#e8e8e8; font-size:12px;}
            QTableWidget::item{padding:4px 6px;}
            QTableWidget::item:alternate{background:#fafafa;}
            QHeaderView::section{background:#fafafa; border:1px solid #e8e8e8;
                padding:5px 8px; font-weight:bold; font-size:12px; color:#333;}
            QLabel{font-size:13px; color:#333;}
        """)

    # ========== 浏览器控制 ==========

    def _find_chrome_path(self):
        """查找本地 Chrome 可执行文件路径"""
        candidates = [
            os.path.expandvars(
                r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(
                r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(
                r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
        for p in candidates:
            if os.path.isfile(p):
                return p
        return None

    def _get_chrome_version(self, chrome_path):
        """获取 Chrome 版本号"""
        import winreg
        try:
            # 从注册表读取
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Google\Chrome\BLBeacon")
            version, _ = winreg.QueryValueEx(key, "version")
            return version
        except Exception:
            pass

        # 从文件版本信息读取
        try:
            import ctypes
            size = ctypes.windll.version.GetFileVersionInfoSizeW(
                chrome_path, None)
            if size:
                buf = ctypes.create_string_buffer(size)
                ctypes.windll.version.GetFileVersionInfoW(
                    chrome_path, None, size, buf)
                res = ctypes.create_string_buffer(size)
                ctypes.windll.version.VerQueryValueW(
                    buf, r"\VarFileInfo\Translation", ctypes.byref(res),
                    ctypes.byref(ctypes.c_ulong()))
                trans = res.raw[:8]
                lang_code = trans.hex()
                query = f"\\StringFileInfo\\{lang_code}\\FileVersion"
                ctypes.windll.version.VerQueryValueW(
                    buf, query, ctypes.byref(res),
                    ctypes.byref(ctypes.c_ulong()))
                return res.value.decode('utf-16le').rstrip('\x00')
        except Exception:
            pass

        return None

    def _download_chromedriver(self, chrome_version):
        """从国内镜像下载 ChromeDriver，自动匹配版本"""
        major = chrome_version.split('.')[0]

        # 缓存目录
        cache_dir = os.path.join(
            os.path.expanduser("~"), ".cache", "cmcc_crawler")
        os.makedirs(cache_dir, exist_ok=True)

        driver_dir = os.path.join(cache_dir, f"chromedriver_{major}")
        driver_exe = os.path.join(driver_dir, "chromedriver.exe")

        # 已存在则直接返回
        if os.path.isfile(driver_exe):
            self.append_log(f"[信息] 使用缓存的 ChromeDriver")
            return driver_exe

        self.append_log(f"[信息] 正在查找 Chrome {major}.x 对应的 ChromeDriver...")

        mirror_base = "https://registry.npmmirror.com/-/binary"

        # 1. 通过 API 查询该大版本号对应的最新 ChromeDriver 版本
        matched_version = None
        try:
            api_url = (
                "https://googlechromelabs.github.io/chrome-for-testing/"
                "latest-patch-versions-per-build-with-downloads.json")
            req = urllib.request.Request(api_url, headers={
                "User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode())

            builds = data.get("builds", {})
            # 查找 major.x 开头的 build
            for build_key in sorted(builds.keys(), reverse=True):
                if build_key.startswith(f"{major}."):
                    build_info = builds[build_key]
                    dl = build_info.get("downloads", {}).get(
                        "chromedriver", [])
                    for d in dl:
                        if d["platform"] == "win64":
                            matched_version = build_info["version"]
                            break
                    if matched_version:
                        break
        except Exception as e:
            self.append_log(f"[警告] API查询失败: {e}")

        # 2. 如果 API 查到了版本，用精确版本下载
        if matched_version:
            self.append_log(
                f"[信息] 匹配到 ChromeDriver 版本: {matched_version}")
            try:
                dl_url = (
                    f"{mirror_base}/chrome-for-testing/"
                    f"{matched_version}/win64/chromedriver-win64.zip")
                self.append_log(f"[信息] 正在下载: {dl_url}")
                zip_path = os.path.join(cache_dir, "chromedriver.zip")
                urllib.request.urlretrieve(dl_url, zip_path)

                with zipfile.ZipFile(zip_path, 'r') as zf:
                    zf.extractall(cache_dir)

                extracted = os.path.join(
                    cache_dir, "chromedriver-win64", "chromedriver.exe")
                if os.path.isfile(extracted):
                    os.makedirs(driver_dir, exist_ok=True)
                    shutil.move(extracted, driver_exe)
                    shutil.rmtree(
                        os.path.join(cache_dir, "chromedriver-win64"),
                        ignore_errors=True)
                os.remove(zip_path, ignore_errors=True)
                self.append_log("[信息] ChromeDriver 下载完成")
                return driver_exe

            except Exception as e:
                self.append_log(f"[警告] 精确版本下载失败: {e}")

        # 3. 回退：尝试直接用 Chrome 精确版本号
        try:
            dl_url = (
                f"{mirror_base}/chrome-for-testing/"
                f"{chrome_version}/win64/chromedriver-win64.zip")
            self.append_log(f"[信息] 尝试精确版本下载: {dl_url}")
            zip_path = os.path.join(cache_dir, "chromedriver2.zip")
            urllib.request.urlretrieve(dl_url, zip_path)

            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(cache_dir)

            extracted = os.path.join(
                cache_dir, "chromedriver-win64", "chromedriver.exe")
            if os.path.isfile(extracted):
                os.makedirs(driver_dir, exist_ok=True)
                shutil.move(extracted, driver_exe)
                shutil.rmtree(
                    os.path.join(cache_dir, "chromedriver-win64"),
                    ignore_errors=True)
            os.remove(zip_path, ignore_errors=True)
            self.append_log("[信息] ChromeDriver 下载完成")
            return driver_exe

        except Exception as e:
            self.append_log(f"[警告] 精确版本下载失败: {e}")

        return None

    def _create_driver(self):
        """创建 Chrome WebDriver"""
        from selenium.webdriver.chrome.service import Service

        chrome_path = self._find_chrome_path()
        if not chrome_path:
            raise Exception(
                "未找到 Chrome 浏览器！请确认已安装 Google Chrome。")

        self.append_log(f"[信息] 找到 Chrome: {chrome_path}")

        # 优先检查程序同目录下是否有 chromedriver.exe
        # 兼容 PyInstaller 打包后的路径
        if getattr(sys, 'frozen', False):
            app_dir = os.path.dirname(sys.executable)
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))
        local_driver = os.path.join(app_dir, "chromedriver.exe")
        if os.path.isfile(local_driver):
            self.append_log(
                f"[信息] 使用本地 ChromeDriver: {local_driver}")
            opts = Options()
            opts.add_argument("--start-maximized")
            opts.add_argument(
                "--disable-blink-features=AutomationControlled")
            opts.add_experimental_option(
                "excludeSwitches", ["enable-automation"])
            opts.add_experimental_option(
                "useAutomationExtension", False)
            service = Service(executable_path=local_driver)
            return webdriver.Chrome(service=service, options=opts)

        # 获取 Chrome 版本
        chrome_version = self._get_chrome_version(chrome_path)
        if chrome_version:
            self.append_log(f"[信息] Chrome 版本: {chrome_version}")
        else:
            chrome_version = ""
            self.append_log("[警告] 无法获取 Chrome 版本，尝试自动匹配")

        # 尝试下载 ChromeDriver
        driver_path = None
        if chrome_version:
            driver_path = self._download_chromedriver(chrome_version)

        opts = Options()
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)

        try:
            if driver_path and os.path.isfile(driver_path):
                service = Service(executable_path=driver_path)
                driver = webdriver.Chrome(service=service, options=opts)
            else:
                # 回退到自动管理
                self.append_log("[信息] 尝试 Selenium 自动管理...")
                driver = webdriver.Chrome(options=opts)

            self.append_log("[信息] Chrome 浏览器已启动")
            return driver

        except Exception as e:
            raise Exception(
                f"启动 Chrome 失败: {e}\n\n"
                "请手动下载 ChromeDriver:\n"
                "1. 打开 Chrome 设置 → 关于 Chrome，记下版本号\n"
                "2. 访问 https://googlechromelabs.github.io/chrome-for-testing/\n"
                "   找到 Stable 版本中与你版本匹配的 chromedriver\n"
                "3. 下载 win64/chromedriver-win64.zip 并解压\n"
                "4. 将 chromedriver.exe 放到程序同目录下\n"
                "   （与 cmcc_fee_crawler.py 放一起）")

    def launch_browser(self):
        if self.driver:
            self.append_log("[警告] 浏览器已在运行中")
            return
        try:
            self.append_log("[信息] 正在启动浏览器...")
            self.update_status("启动浏览器中...")

            self.driver = self._create_driver()

            # 反检测
            self.driver.execute_cdp_cmd(
                "Page.addScriptToEvaluateOnNewDocument", {"source": """
                    Object.defineProperty(navigator,'webdriver',{
                        get:()=>undefined});
                """})

            self.driver.get(TARGET_URL)
            self.append_log("[信息] 浏览器已启动，已打开资费公示页面")
            self.append_log(
                "[提示] 请在浏览器中手动选择省份（如江西省）")
            self.append_log(
                "[提示] 选择完毕后，点击「开始抓取」按钮")
            self.append_log(
                "[提示] 程序将自动遍历所有资费类型并抓取数据")

            self.btn_launch.setEnabled(False)
            self.btn_start.setEnabled(True)
            self.update_status("等待用户操作浏览器...")

        except Exception as e:
            self.append_log(f"[错误] 启动浏览器失败: {str(e)}")
            self.update_status("浏览器启动失败")
            self.driver = None
            self.btn_launch.setEnabled(True)

    def start_crawl(self):
        if not self.driver:
            self.append_log("[错误] 请先启动浏览器")
            return
        try:
            _ = self.driver.current_url
        except Exception:
            self.append_log("[错误] 浏览器已关闭，请重新启动")
            self._reset_browser_state()
            return

        self.is_crawling = True
        self.stop_event.clear()
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_launch.setEnabled(False)

        # 默认遍历所有资费类型
        self.crawler_thread = CrawlerThread(
            self.driver, self.signals, self.stop_event,
            traverse_types=True)
        self.crawler_thread.start()

    def stop_crawl(self):
        self.append_log("[信息] 正在停止抓取...")
        self.stop_event.set()
        self.btn_stop.setEnabled(False)

    # ========== 数据接收（主线程，只做 UI 更新） ==========

    def _on_new_data(self, batch):
        """收到爬虫线程发来的新数据，追加到表格"""
        if not batch:
            return
        self.all_data.extend(batch)
        self.append_log(f"[信息] 新增 {len(batch)} 条，"
                        f"累计 {len(self.all_data)} 条")
        # 增量更新表格（只追加新行，不全量刷新）
        self._append_rows(batch)

    def _append_rows(self, rows):
        """增量追加行到表格（比全量刷新快得多）"""
        # 如果有筛选条件，走全量刷新
        if (self.search_input.text().strip()
                or self.price_min.text().strip()
                or self.price_max.text().strip()):
            self.apply_filter()
            return

        # 无筛选条件，直接追加
        row_count = self.table.rowCount()
        self.table.setRowCount(row_count + len(rows))
        for i, rd in enumerate(rows):
            for j, v in enumerate(rd):
                item = QTableWidgetItem(str(v) if v else "")
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignLeft
                    | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row_count + i, j, item)

        self.label_count.setText(f"共 {len(self.all_data)} 条数据")

    # ========== 筛选 ==========

    def apply_filter(self):
        keyword = self.search_input.text().strip().lower()
        pmin = self.price_min.text().strip()
        pmax = self.price_max.text().strip()
        price_lo = float(pmin) if pmin else 0
        price_hi = float(pmax) if pmax else float('inf')

        filtered = []
        for row in self.all_data:
            if keyword:
                if not (keyword in row[COL_PLAN_ID].lower()
                        or keyword in row[COL_NAME].lower()):
                    continue
            price = _parse_price(row[COL_PRICE])
            if price is not None and not (price_lo <= price <= price_hi):
                continue
            filtered.append(row)

        self.table.setRowCount(len(filtered))
        for i, rd in enumerate(filtered):
            for j, v in enumerate(rd):
                item = QTableWidgetItem(str(v) if v else "")
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignLeft
                    | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(i, j, item)

        total = len(self.all_data)
        shown = len(filtered)
        self.label_count.setText(
            f"共 {total} 条数据"
            + (f"（筛选显示 {shown} 条）" if total != shown else ""))

    def reset_filter(self):
        self.search_input.clear()
        self.price_min.clear()
        self.price_max.clear()
        self.apply_filter()

    # ========== 数据操作 ==========

    def clear_data(self):
        if self.is_crawling:
            self.append_log("[警告] 请先停止抓取再清空数据")
            return
        if QMessageBox.question(
            self, "确认清空", "确定要清空所有已抓取的数据吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            self.all_data.clear()
            self.seen_ids.clear()
            self.table.setRowCount(0)
            self.label_count.setText("共 0 条数据")
            self.append_log("[信息] 数据已清空")

    def export_excel(self):
        if not self.all_data:
            QMessageBox.warning(self, "提示", "没有数据可导出")
            return
        if not HAS_OPENPYXL:
            QMessageBox.critical(
                self, "错误", "缺少 openpyxl，请执行: pip install openpyxl")
            return

        # 获取筛选后的数据
        keyword = self.search_input.text().strip().lower()
        pmin = self.price_min.text().strip()
        pmax = self.price_max.text().strip()
        price_lo = float(pmin) if pmin else 0
        price_hi = float(pmax) if pmax else float('inf')

        export = []
        for row in self.all_data:
            if keyword:
                if not (keyword in row[COL_PLAN_ID].lower()
                        or keyword in row[COL_NAME].lower()):
                    continue
            price = _parse_price(row[COL_PRICE])
            if price is not None and not (price_lo <= price <= price_hi):
                continue
            export.append(row)

        if not export:
            QMessageBox.warning(self, "提示", "筛选后没有数据可导出")
            return

        default_name = (
            f"中国移动资费数据_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", default_name, "Excel文件 (*.xlsx)")
        if not path:
            return

        try:
            self.append_log(f"[信息] 正在导出 {len(export)} 条数据...")
            _write_excel(path, export)
            self.append_log(f"[信息] 导出成功: {path}")
            QMessageBox.information(
                self, "导出成功",
                f"已导出 {len(export)} 条数据到:\n{path}")
        except Exception as e:
            self.append_log(f"[错误] 导出失败: {str(e)}")
            QMessageBox.critical(self, "导出失败", str(e))

    # ========== UI 辅助 ==========

    def append_log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{ts}] {msg}")
        sb = self.log_text.verticalScrollBar()
        sb.setValue(sb.maximum())

    def update_status(self, text):
        self.label_status.setText(f"状态: {text}")
        if "抓取中" in text:
            color = "#1890ff"
        elif "完成" in text or "成功" in text:
            color = "#52c41a"
        elif "错误" in text or "失败" in text:
            color = "#ff4d4f"
        else:
            color = "#faad14"
        self.label_status.setStyleSheet(
            f"color:{color}; font-weight:bold;")

    def _on_type_switched(self, info):
        """资费类型切换时的回调"""
        self.label_current_type.setText(f"当前: {info}")

    def _on_error(self, msg):
        self.append_log(msg)
        self.update_status("出错")

    def _on_crawl_finished(self):
        self.is_crawling = False
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.append_log(
            f"[信息] 抓取结束，共获取 {len(self.all_data)} 条数据")

    def _reset_browser_state(self):
        self.driver = None
        self.btn_launch.setEnabled(True)
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(False)
        self.is_crawling = False
        self.update_status("浏览器未启动")

    # ========== 关闭 ==========

    def closeEvent(self, event):
        if self.is_crawling:
            if QMessageBox.question(
                self, "确认退出", "正在抓取数据，确定要退出吗？",
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            ) == QMessageBox.StandardButton.No:
                event.ignore()
                return
            self.stop_event.set()

        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None

        event.accept()


# ============================================================
# 工具函数
# ============================================================
def _parse_price(price_str):
    if not price_str:
        return None
    m = re.search(r'([\d.]+)', price_str)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _write_excel(path, data):
    """将数据写入 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资费数据"

    # 表头样式
    h_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill(start_color="1890FF", end_color="1890FF",
                         fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center",
                        wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'))

    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = thin

    # 高亮 4 个新增字段列（第 12-15 列）
    hl_fill = PatternFill(start_color="36CFC9", end_color="36CFC9",
                          fill_type="solid")
    for c in (12, 13, 14, 15):
        ws.cell(row=1, column=c).fill = hl_fill

    # 数据
    d_font = Font(name="微软雅黑", size=10)
    d_align = Alignment(horizontal="left", vertical="center",
                        wrap_text=True)
    alt_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA",
                           fill_type="solid")

    for ri, row in enumerate(data, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci,
                           value=str(val) if val else "")
            cell.font = d_font
            cell.alignment = d_align
            cell.border = thin
            if ri % 2 == 0:
                cell.fill = alt_fill

    # 列宽
    widths = {
        1: 14, 2: 28, 3: 14, 4: 10, 5: 10, 6: 14, 7: 16,
        8: 14, 9: 14, 10: 20, 11: 10, 12: 14, 13: 16,
        14: 14, 15: 10, 16: 22, 17: 22, 18: 28,
    }
    for c, w in widths.items():
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


# ============================================================
# 入口
# ============================================================
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setFont(QFont("Microsoft YaHei", 9))
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
