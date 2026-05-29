# -*- coding: utf-8 -*-
"""
Excel 导出工具
"""

import re

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..constants import COLUMNS


def parse_price(price_str):
    """从价格字符串中提取数值"""
    if not price_str:
        return None
    m = re.search(r'([\d.]+)', price_str)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def write_excel(path, data):
    """将数据写入 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资费数据"

    # 表头样式
    h_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill(start_color="1890FF", end_color="1890FF",
                         fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center",
                        wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'))

    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = thin

    # 高亮 4 个新增字段列（第 12-15 列）
    hl_fill = PatternFill(start_color="36CFC9", end_color="36CFC9",
                          fill_type="solid")
    for c in (12, 13, 14, 15):
        ws.cell(row=1, column=c).fill = hl_fill

    # 数据
    d_font = Font(name="微软雅黑", size=10)
    d_align = Alignment(horizontal="left", vertical="center",
                        wrap_text=True)
    alt_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA",
                           fill_type="solid")

    for ri, row in enumerate(data, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci,
                           value=str(val) if val else "")
            cell.font = d_font
            cell.alignment = d_align
            cell.border = thin
            if ri % 2 == 0:
                cell.fill = alt_fill

    # 列宽
    widths = {
        1: 14, 2: 28, 3: 14, 4: 10, 5: 10, 6: 14, 7: 16,
        8: 14, 9: 14, 10: 20, 11: 10, 12: 14, 13: 16,
        14: 14, 15: 10, 16: 22, 17: 22, 18: 28,
    }
    for c, w in widths.items():
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)